# https://myjamong.tistory.com/53 여기 참고

import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

# create table test(num int (11), name varchar(10));


class Test:
    def __init__(self,num,name):
        self.num = num
        self.name = name

def select_all():
    conn = pymysql.connect(host = 'localhost',user = 'root',password ='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()
            for row in rs:
                print(row)
    finally:
        conn.close()

def insert_test(test_obj):
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values(%s, %s)'
            curs.execute(sql, (test_obj.num, test_obj.name))
        conn.commit()
    finally:
        conn.close()

def delete_test(num):
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'delete from test where num=%s'
            curs.execute(sql, num)
        conn.commit()
    finally:
        conn.close()

def delete_all():
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'delete from test'
            curs.execute(sql)
        conn.commit()
    finally:
        conn.close()


def update_test(test_obj):
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'update test set name=%s where num=%s'
            curs.execute(sql, (test_obj.name, test_obj.num))
        conn.commit()
    finally:
        conn.close()

def select_all_to_excel():
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()

            wb = Workbook()
            ws = wb.active

            #첫행 입력
            ws.append(('번호','이름'))

            #DB 모든 데이터 엑셀로
            for row in rs:
                ws.append(row)

            wb.save('./python_test_excel.xlsx')
    finally:
        conn.close()
        wb.close()
    
    
# if __name__ == "__main__":
#     #데이터 1000개정도 넣기
#     for i in range(1,1000):
#         test = Test(i, str(i) + '이름')
#         insert_test(test)
    
#     #DB -> 엑셀파일
#     select_all_to_excel()

def insert_excel_to_db():
    conn = pymysql.connect(host='localhost', user='root', password='root', db='recipe_db', charset='utf8')
    try:
        with conn.cursor() as curs:
            sql = 'insert into recipe(sub_id,recipe_name,recipe,ingredient,ing_broth,ing_seasoning) values(%s, %s, %s, %s, %s, %s)'
 
            wb = load_workbook('./레시피.xlsx',data_only=True)
            ws = wb['sheet1']
 
            iter_rows = iter(ws.rows)
            next(iter_rows)
            for row in iter_rows:
                curs.execute(sql, (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value))
            conn.commit()
    finally:
        conn.close()
        wb.close()
 
 
if __name__ == "__main__":
    # delete_all()
    insert_excel_to_db()
    # select_all()



